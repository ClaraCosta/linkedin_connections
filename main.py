from __future__ import annotations

import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

import settings
from consts import (
    CONNECT_LABEL,
    LINKEDIN_GROW_URL,
    LINKEDIN_HOME_URL,
    LINKEDIN_MY_NETWORK_URL,
    MAX_SCROLL_ATTEMPTS_WITHOUT_CLICK,
    MY_NETWORK_LABEL,
    PENDING_LABEL,
)


@dataclass
class RunStats:
    clicked: int = 0
    skipped: int = 0
    scrolls_without_click: int = 0


@dataclass
class PersonInfo:
    name: str
    description: str
    profile_url: str


def build_driver() -> WebDriver:
    Path(settings.CHROMEDRIVER_LOG_PATH).parent.mkdir(parents=True, exist_ok=True)

    options = Options()
    options.binary_location = settings.CHROME_BINARY

    if settings.ATTACH_TO_EXISTING_CHROME:
        print(f"Anexando ao Chrome ja aberto em {settings.CHROME_DEBUGGER_ADDRESS}...")
        ensure_debugger_is_available()
        options.debugger_address = settings.CHROME_DEBUGGER_ADDRESS
        return webdriver.Chrome(
            service=Service(log_output=settings.CHROMEDRIVER_LOG_PATH),
            options=options,
        )

    Path(settings.CHROME_USER_DATA_DIR).mkdir(parents=True, exist_ok=True)
    print(f"Abrindo novo Chrome com perfil em {settings.CHROME_USER_DATA_DIR}...")
    options.add_argument(f"--user-data-dir={settings.CHROME_USER_DATA_DIR}")
    options.add_argument(f"--profile-directory={settings.CHROME_PROFILE_DIRECTORY}")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--remote-debugging-port=0")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    if settings.START_MAXIMIZED:
        options.add_argument("--start-maximized")

    # Selenium Manager downloads or finds a compatible driver automatically.
    return webdriver.Chrome(
        service=Service(log_output=settings.CHROMEDRIVER_LOG_PATH),
        options=options,
    )


def ensure_debugger_is_available() -> None:
    url = f"http://{settings.CHROME_DEBUGGER_ADDRESS}/json/version"
    try:
        with urllib.request.urlopen(url, timeout=2) as response:
            if response.status == 200:
                return
    except urllib.error.URLError as exc:
        raise RuntimeError(
            "Nao consegui anexar ao Chrome aberto. "
            "Abra o Chrome Default com depuracao remota antes de rodar o RPA. "
            "Use: ./start_chrome_default_debug.sh"
        ) from exc


def wait_for_page_ready(driver: WebDriver) -> None:
    WebDriverWait(driver, settings.WAIT_SECONDS).until(
        lambda current_driver: current_driver.execute_script(
            "return document.readyState"
        )
        == "complete"
    )


def visible_text(element: WebElement) -> str:
    try:
        return (element.text or "").strip()
    except StaleElementReferenceException:
        return ""


def aria_label(element: WebElement) -> str:
    try:
        return (element.get_attribute("aria-label") or "").strip()
    except StaleElementReferenceException:
        return ""


def has_label(element: WebElement, candidates: Iterable[str]) -> bool:
    labels = tuple(candidate.lower() for candidate in candidates)
    text = visible_text(element).lower()
    aria = aria_label(element).lower()
    return any(label in text or label in aria for label in labels)


def open_my_network(driver: WebDriver) -> None:
    print("Abrindo LinkedIn...")
    driver.get(LINKEDIN_HOME_URL)
    wait_for_page_ready(driver)

    if "/login" in driver.current_url:
        print(
            "Login do LinkedIn necessario. Entre manualmente no Chrome aberto e rode novamente.",
            file=sys.stderr,
        )
        raise SystemExit(2)

    wait = WebDriverWait(driver, settings.WAIT_SECONDS)
    try:
        print("Clicando em Minha rede...")
        menu_item = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    (
                        "//a[contains(translate(@aria-label,"
                        " 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ',"
                        " 'abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'),"
                        f" '{MY_NETWORK_LABEL.lower()}')"
                        " or .//*[normalize-space()='Minha rede' or normalize-space()='Minha Rede']]"
                    ),
                )
            )
        )
        menu_item.click()
        wait_for_page_ready(driver)
    except TimeoutException:
        print("Nao encontrei o menu Minha rede; abrindo a URL diretamente...")
        driver.get(LINKEDIN_MY_NETWORK_URL)
        wait_for_page_ready(driver)


def go_to_people_suggestions(driver: WebDriver) -> None:
    print("Abrindo sugestoes de conexao...")
    driver.get(LINKEDIN_GROW_URL)
    wait_for_page_ready(driver)


def find_connect_buttons(driver: WebDriver) -> list[WebElement]:
    xpath = (
        "//button[not(@disabled) and "
        "(contains(normalize-space(.), 'Conectar') or contains(@aria-label, 'Conectar'))]"
        " | "
        "//a[@aria-disabled='false' and "
        "(contains(normalize-space(.), 'Conectar') or contains(@aria-label, 'Conectar'))]"
    )

    buttons = driver.find_elements(By.XPATH, xpath)
    visible_buttons: list[WebElement] = []

    for button in buttons:
        try:
            if button.is_displayed() and has_label(button, [CONNECT_LABEL]):
                visible_buttons.append(button)
        except StaleElementReferenceException:
            continue

    return visible_buttons


def scroll_button_into_view(driver: WebDriver, button: WebElement) -> None:
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
        button,
    )
    time.sleep(0.4)


def safe_click(driver: WebDriver, button: WebElement) -> bool:
    label_before_click = aria_label(button) or visible_text(button)
    person_hint = extract_person_hint(label_before_click)

    if settings.DRY_RUN:
        print(f"[dry-run] Botao encontrado: {label_before_click}")
        return True

    try:
        scroll_button_into_view(driver, button)
        button.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", button)
    except (StaleElementReferenceException, WebDriverException) as exc:
        print(f"Pulando botao instavel: {exc.__class__.__name__}")
        return False

    handle_invite_modal_if_needed(driver)
    return wait_until_pending(driver, button, person_hint)


def get_person_info(driver: WebDriver, button: WebElement) -> PersonInfo:
    label = aria_label(button) or visible_text(button)
    name_from_label = extract_person_hint(label)

    try:
        card_html, card_text = driver.execute_script(
            """
            const button = arguments[0];
            let node = button;

            for (let depth = 0; node && depth < 10; depth += 1) {
              const text = (node.innerText || '').trim();
              const profileLink = node.querySelector
                ? node.querySelector('a[href*="/in/"]')
                : null;
              const paragraphs = node.querySelectorAll
                ? Array.from(node.querySelectorAll('p'))
                : [];
              const connectActions = node.querySelectorAll
                ? Array.from(node.querySelectorAll('button,a'))
                    .filter((element) => {
                      const label = `${element.innerText || ''} ${element.getAttribute('aria-label') || ''}`;
                      return label.includes('Conectar');
                    })
                : [];

              if (
                text.includes('Conectar')
                && paragraphs.length >= 2
                && connectActions.length <= 2
                && (profileLink || text.length > 20)
              ) {
                return [node.outerHTML || '', text];
              }

              node = node.parentElement;
            }

            return ['', (button.innerText || button.getAttribute('aria-label') || '').trim()];
            """,
            button,
        )
    except WebDriverException:
        card_html = ""
        card_text = label

    person = parse_person_info_with_beautifulsoup(str(card_html), name_from_label)

    if person.name != "Nome nao identificado" or person.description:
        return person

    lines = normalize_card_lines(str(card_text))
    name = name_from_label or first_meaningful_line(lines) or person.name
    description = first_description_line(lines, name)

    return PersonInfo(
        name=name,
        description=description,
        profile_url=person.profile_url,
    )


def parse_person_info_with_beautifulsoup(
    card_html: str,
    name_from_label: str,
) -> PersonInfo:
    if not card_html:
        return PersonInfo(
            name=name_from_label or "Nome nao identificado",
            description="",
            profile_url="",
        )

    soup = BeautifulSoup(card_html, "html.parser")
    profile_link = soup.find("a", href=lambda href: href and "/in/" in href)
    profile_url = profile_link.get("href", "") if profile_link else ""

    paragraph_lines = normalize_card_lines(
        "\n".join(paragraph.get_text(" ", strip=True) for paragraph in soup.find_all("p"))
    )
    all_lines = normalize_card_lines(soup.get_text("\n", strip=True))

    name = name_from_label or first_meaningful_line(paragraph_lines) or first_meaningful_line(all_lines)
    name = name or "Nome nao identificado"
    description = first_description_line(paragraph_lines, name)

    if not description:
        description = first_description_line(all_lines, name)

    return PersonInfo(
        name=name,
        description=description,
        profile_url=profile_url,
    )


def normalize_card_lines(text: str) -> list[str]:
    ignored = {
        "",
        CONNECT_LABEL,
        PENDING_LABEL,
        "Seguir",
        "Enviar mensagem",
        "Mensagem",
        "Mais",
    }
    lines: list[str] = []

    for raw_line in text.splitlines():
        line = " ".join(raw_line.split())
        line = line.replace(" Verificado", "").strip()
        if line not in ignored and line not in lines:
            lines.append(line)

    return lines


def is_noise_line(line: str, name: str = "") -> bool:
    lowered = line.lower()
    normalized_name = name.strip().lower()

    if not line:
        return True
    if normalized_name and (
        lowered == normalized_name
        or lowered.startswith(f"{normalized_name},")
        or lowered.startswith(f"{normalized_name} ")
    ):
        return True
    if normalized_name and normalized_name in lowered and (
        "premium" in lowered
        or "verificado" in lowered
        or "verified" in lowered
    ):
        return True
    if lowered.startswith(("conectar", "pendente", "seguir", "patrocinado")):
        return True
    if lowered in {"premium", "verificado", "verified"}:
        return True
    if lowered in {"1º", "2º", "3º", "1°", "2°", "3°"}:
        return True
    if "conex" in lowered and (
        "em comum" in lowered
        or "mútua" in lowered
        or "mutua" in lowered
        or "conexão" in lowered
        or "conexões" in lowered
        or "conexao" in lowered
        or "conexoes" in lowered
    ):
        return True

    return False


def first_meaningful_line(lines: list[str]) -> str:
    for line in lines:
        if is_noise_line(line):
            continue

        return line

    return ""


def first_description_line(lines: list[str], name: str) -> str:
    for line in lines:
        if is_noise_line(line, name):
            continue

        return line

    return ""


def save_person_to_xlsx(person: PersonInfo, status: str) -> None:
    output_path = Path(settings.OUTPUT_XLSX_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = ["data_hora", "nome", "descricao", "status", "perfil_linkedin"]

    if output_path.exists():
        workbook = load_workbook(output_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "conexoes"
        sheet.append(headers)

    sheet.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            person.name,
            person.description,
            status,
            person.profile_url,
        ]
    )
    workbook.save(output_path)


def extract_person_hint(label: str) -> str:
    cleaned = label.strip()
    prefixes = [
        "Convidar ",
        "Conectar-se a ",
        "Conectar com ",
        "Conectar a ",
        "Conectar ",
    ]

    for prefix in prefixes:
        if cleaned.startswith(prefix):
            cleaned = cleaned.removeprefix(prefix).strip()
            break

    suffixes = [
        " para se conectar",
        " para conectar",
    ]

    for suffix in suffixes:
        if cleaned.endswith(suffix):
            return cleaned.removesuffix(suffix).strip()

    return cleaned if cleaned != label.strip() else ""


def status_from_click_result(clicked: bool) -> str:
    if settings.DRY_RUN:
        return "dry_run"

    return "pendente_confirmado" if clicked else "nao_confirmado"


def wait_until_pending(
    driver: WebDriver,
    original_button: WebElement,
    person_hint: str,
) -> bool:
    def changed_to_pending(_: WebDriver) -> bool:
        try:
            if has_label(original_button, [PENDING_LABEL]):
                return True
        except StaleElementReferenceException:
            pass

        if person_hint:
            pending_for_person_xpath = (
                "//*[self::button or self::a]"
                "[contains(@aria-label, 'Pendente') and "
                f"contains(@aria-label, {xpath_literal(person_hint)})]"
            )
            return any(
                element.is_displayed()
                for element in driver.find_elements(By.XPATH, pending_for_person_xpath)
            )

        pending_xpath = (
            "//*[self::button or self::a]"
            "[contains(@aria-label, 'Pendente') or contains(normalize-space(.), 'Pendente')]"
        )
        return any(
            element.is_displayed()
            for element in driver.find_elements(By.XPATH, pending_xpath)
        )

    try:
        WebDriverWait(driver, settings.WAIT_SECONDS).until(changed_to_pending)
        return True
    except TimeoutException:
        print("Clique feito, mas nao consegui confirmar estado Pendente.")
        return False


def xpath_literal(value: str) -> str:
    if "'" not in value:
        return f"'{value}'"

    if '"' not in value:
        return f'"{value}"'

    parts = value.split("'")
    return "concat(" + ', "\'", '.join(f"'{part}'" for part in parts) + ")"


def handle_invite_modal_if_needed(driver: WebDriver) -> None:
    modal_xpaths = [
        "//button[contains(@aria-label, 'Enviar sem nota') or normalize-space()='Enviar sem nota']",
        "//button[contains(@aria-label, 'Enviar') or normalize-space()='Enviar']",
        "//button[contains(@aria-label, 'Enviar agora') or normalize-space()='Enviar agora']",
        "//button[contains(@aria-label, 'Fechar') or normalize-space()='Fechar']",
    ]

    for xpath in modal_xpaths:
        try:
            button = driver.find_element(By.XPATH, xpath)
            if button.is_displayed() and button.is_enabled():
                button.click()
                time.sleep(0.5)
                return
        except NoSuchElementException:
            continue
        except WebDriverException:
            continue


def connect_with_people(driver: WebDriver) -> RunStats:
    stats = RunStats()

    while stats.clicked < settings.DAILY_CONNECTION_LIMIT:
        buttons = find_connect_buttons(driver)

        if not buttons:
            stats.scrolls_without_click += 1
            if stats.scrolls_without_click >= MAX_SCROLL_ATTEMPTS_WITHOUT_CLICK:
                break

            driver.execute_script("window.scrollBy(0, Math.floor(window.innerHeight * 0.85));")
            time.sleep(settings.SCROLL_PAUSE_SECONDS)
            continue

        stats.scrolls_without_click = 0
        button = buttons[0]
        person = get_person_info(driver, button)
        print(f"Tentando conectar: {person.name} - {person.description}")

        clicked = safe_click(driver, button)
        save_person_to_xlsx(person, status_from_click_result(clicked))

        if clicked:
            stats.clicked += 1
            print(f"Convites confirmados: {stats.clicked}/{settings.DAILY_CONNECTION_LIMIT}")
        else:
            stats.skipped += 1

        time.sleep(settings.CLICK_PAUSE_SECONDS)

    return stats


def main() -> int:
    driver: WebDriver | None = None

    try:
        driver = build_driver()
        open_my_network(driver)
        go_to_people_suggestions(driver)
        stats = connect_with_people(driver)
        print(
            "Finalizado: "
            f"{stats.clicked} convite(s) enviado(s), "
            f"{stats.skipped} botao(oes) pulado(s)."
        )
        return 0
    except WebDriverException as exc:
        print(
            "Falha ao iniciar/controlar o Chrome. "
            "Veja o log em "
            f"{settings.CHROMEDRIVER_LOG_PATH}. "
            "Se voce configurou o perfil diario e ele ja estiver aberto, "
            "feche todas as janelas do Chrome ou ajuste CHROME_USER_DATA_DIR/"
            "CHROME_PROFILE_DIRECTORY em settings.py.",
            file=sys.stderr,
        )
        print(str(exc), file=sys.stderr)
        return 1
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    finally:
        if driver is not None and not settings.KEEP_BROWSER_OPEN:
            driver.quit()


if __name__ == "__main__":
    raise SystemExit(main())
